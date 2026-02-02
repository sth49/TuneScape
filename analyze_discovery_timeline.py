"""
Discovery Timeline Analysis

Analyzes which parameter changes led to marginal coverage gains over time.
Shows the sequence of discoveries and what caused them.
"""
import ast
import json
from pathlib import Path
from collections import defaultdict
import numpy as np
import openpyxl

DATA_DIR = Path("data_VIS26/data_VIS26")
TUNERS = ["SymTuner", "CMA_ES", "Genetic", "SuccessiveHalving"]
PROGRAMS = ["gawk", "gcal", "grep"]

EXCLUDE_FROM_PARAMS = ["Iteration Coverage", "Accumulative Coverage"]


def load_data(program, tuner):
    """Load parameters and coverage data."""
    base_path = DATA_DIR / program / tuner
    xlsx_path = base_path / "parameters.xlsx"
    coverage_path = base_path / "coverage_set"

    if not xlsx_path.exists() or not coverage_path.exists():
        return None

    # Load parameters
    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb.active

    param_names = []
    param_rows = []
    for row in range(3, sheet.max_row + 1):
        val = sheet.cell(row, 1).value
        if val and val not in EXCLUDE_FROM_PARAMS:
            param_names.append(val)
            param_rows.append(row)

    n_trials = sheet.max_column - 1
    trials_params = []

    for col in range(2, sheet.max_column + 1):
        params = {}
        for param_name, row in zip(param_names, param_rows):
            val = sheet.cell(row, col).value
            if val is True:
                params[param_name] = True
            elif val is False:
                params[param_name] = False
            elif val is not None:
                try:
                    params[param_name] = float(val) if isinstance(val, (int, float)) else val
                except:
                    params[param_name] = str(val)
            else:
                params[param_name] = None
        trials_params.append(params)

    wb.close()

    # Load coverage sets
    coverage_sets = []
    with open(coverage_path, 'r') as f:
        for line in f:
            line = line.strip()
            if line:
                coverage_set = ast.literal_eval(line)
                coverage_sets.append(set(coverage_set))

    # Align lengths
    min_len = min(len(trials_params), len(coverage_sets))
    trials_params = trials_params[:min_len]
    coverage_sets = coverage_sets[:min_len]

    # Calculate marginal coverage
    seen_branches = set()
    trials = []

    for i, (params, cov_set) in enumerate(zip(trials_params, coverage_sets)):
        new_branches = cov_set - seen_branches
        marginal = len(new_branches)
        cumulative = len(seen_branches) + marginal
        seen_branches.update(cov_set)

        trials.append({
            "trial_id": i + 1,
            "parameters": params,
            "coverage": len(cov_set),
            "marginal_coverage": marginal,
            "cumulative_coverage": cumulative,
            "new_branches": list(new_branches)[:20]  # Keep first 20 for reference
        })

    return {
        "param_names": param_names,
        "trials": trials,
        "total_branches": len(seen_branches)
    }


def find_param_changes(prev_params, curr_params, param_names):
    """Find which parameters changed between two trials."""
    changes = []
    for param in param_names:
        prev_val = prev_params.get(param)
        curr_val = curr_params.get(param)

        if prev_val != curr_val:
            changes.append({
                "param": param,
                "from": prev_val,
                "to": curr_val
            })

    return changes


def analyze_discoveries(data):
    """Analyze parameter changes that led to discoveries."""
    trials = data["trials"]
    param_names = data["param_names"]

    discoveries = []
    last_discovery_params = None
    last_discovery_idx = -1

    # Track parameter value -> coverage statistics
    param_value_stats = defaultdict(lambda: {"coverages": [], "marginals": []})

    for i, trial in enumerate(trials):
        # Track stats for all parameter values
        for param in param_names:
            val = trial["parameters"].get(param)
            if val is not None:
                key = (param, str(val))
                param_value_stats[key]["coverages"].append(trial["coverage"])
                if trial["marginal_coverage"] > 0:
                    param_value_stats[key]["marginals"].append(trial["marginal_coverage"])

        # Only track trials with marginal coverage
        if trial["marginal_coverage"] > 0:
            if last_discovery_params is not None:
                # Find what changed since last discovery
                changes = find_param_changes(
                    last_discovery_params,
                    trial["parameters"],
                    param_names
                )
            else:
                # First discovery - no previous to compare
                changes = []

            discoveries.append({
                "trial_id": trial["trial_id"],
                "marginal_coverage": trial["marginal_coverage"],
                "cumulative_coverage": trial["cumulative_coverage"],
                "total_coverage": trial["coverage"],
                "param_changes": changes,
                "n_changes": len(changes),
                "trials_since_last": trial["trial_id"] - last_discovery_idx - 1 if last_discovery_idx >= 0 else 0,
                "parameters": trial["parameters"]
            })

            last_discovery_params = trial["parameters"].copy()
            last_discovery_idx = i

    # Calculate parameter value statistics
    param_effects = []
    for (param, val), stats in param_value_stats.items():
        if len(stats["coverages"]) >= 5:  # At least 5 samples
            param_effects.append({
                "param": param,
                "value": val,
                "avg_coverage": np.mean(stats["coverages"]),
                "std_coverage": np.std(stats["coverages"]),
                "n_trials": len(stats["coverages"]),
                "n_discoveries": len(stats["marginals"]),
                "total_marginal": sum(stats["marginals"]),
                "discovery_rate": len(stats["marginals"]) / len(stats["coverages"])
            })

    # Sort by discovery rate
    param_effects.sort(key=lambda x: x["discovery_rate"], reverse=True)

    return discoveries, param_effects


def build_discovery_tree(discoveries, max_depth=3):
    """Build a tree structure showing discovery paths."""
    if not discoveries:
        return None

    # Find most common parameter changes
    param_change_counts = defaultdict(int)
    param_change_marginals = defaultdict(int)

    for disc in discoveries:
        for change in disc["param_changes"]:
            key = change["param"]
            param_change_counts[key] += 1
            param_change_marginals[key] += disc["marginal_coverage"]

    # Sort by total marginal contribution
    sorted_params = sorted(
        param_change_marginals.items(),
        key=lambda x: x[1],
        reverse=True
    )

    # Build simplified tree
    tree = {
        "type": "root",
        "trial_id": 0,
        "cumulative": 0,
        "children": []
    }

    # Group discoveries by their first significant param change
    for disc in discoveries[:50]:  # Limit to first 50 discoveries
        if disc["param_changes"]:
            # Find the most impactful change (by param importance)
            best_change = None
            best_rank = float('inf')

            for change in disc["param_changes"]:
                for rank, (param, _) in enumerate(sorted_params):
                    if param == change["param"] and rank < best_rank:
                        best_change = change
                        best_rank = rank
                        break

            if best_change:
                tree["children"].append({
                    "type": "discovery",
                    "trial_id": disc["trial_id"],
                    "marginal": disc["marginal_coverage"],
                    "cumulative": disc["cumulative_coverage"],
                    "key_change": best_change,
                    "all_changes": disc["param_changes"][:5],  # Top 5 changes
                    "trials_since_last": disc["trials_since_last"]
                })
        else:
            # First discovery or no changes
            tree["children"].append({
                "type": "discovery",
                "trial_id": disc["trial_id"],
                "marginal": disc["marginal_coverage"],
                "cumulative": disc["cumulative_coverage"],
                "key_change": None,
                "all_changes": [],
                "trials_since_last": disc["trials_since_last"]
            })

    return tree


def main():
    results = {}

    for program in PROGRAMS:
        results[program] = {}
        print(f"\n{'='*50}")
        print(f"Program: {program}")
        print('='*50)

        for tuner in TUNERS:
            print(f"\n  {tuner}:")
            data = load_data(program, tuner)

            if data is None:
                print("    No data found")
                continue

            discoveries, param_effects = analyze_discoveries(data)
            tree = build_discovery_tree(discoveries)

            print(f"    Trials: {len(data['trials'])}")
            print(f"    Discoveries: {len(discoveries)}")
            print(f"    Total branches: {data['total_branches']}")

            if discoveries:
                # Top parameters by discovery contribution
                param_contribution = defaultdict(int)
                for disc in discoveries:
                    for change in disc["param_changes"]:
                        param_contribution[change["param"]] += disc["marginal_coverage"]

                top_params = sorted(param_contribution.items(), key=lambda x: x[1], reverse=True)[:5]
                print(f"    Top contributing params: {[p[0] for p in top_params]}")

            # Prepare all trials data (simplified, without param_changes for size)
            all_trials = []
            for t in data["trials"]:  # All trials
                all_trials.append({
                    "trial_id": t["trial_id"],
                    "parameters": t["parameters"],
                    "coverage": t["coverage"],
                    "marginal_coverage": t["marginal_coverage"],
                    "cumulative_coverage": t["cumulative_coverage"],
                })

            results[program][tuner] = {
                "n_trials": len(data["trials"]),
                "n_discoveries": len(discoveries),
                "total_branches": data["total_branches"],
                "discoveries": discoveries[:100],  # Limit to 100
                "all_trials": all_trials,  # All trials for tree view
                "param_effects": param_effects[:50],  # Top 50
                "discovery_tree": tree,
                "summary": {
                    "avg_marginal": np.mean([d["marginal_coverage"] for d in discoveries]) if discoveries else 0,
                    "avg_trials_between": np.mean([d["trials_since_last"] for d in discoveries if d["trials_since_last"] > 0]) if discoveries else 0,
                    "discovery_rate": len(discoveries) / len(data["trials"]) if data["trials"] else 0
                }
            }

    # Save results
    output_path = Path("public/data/discovery_timeline_data.json")
    with open(output_path, 'w') as f:
        json.dump(results, f)

    print(f"\n\nSaved to {output_path}")
    print(f"File size: {output_path.stat().st_size / 1024:.1f} KB")


if __name__ == "__main__":
    main()
