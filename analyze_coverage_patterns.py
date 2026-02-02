"""
Coverage Pattern Analysis using PCA + SHAP

Instead of predicting scalar coverage count, we:
1. Create coverage vectors (which branches were covered)
2. Apply PCA to reduce dimensionality
3. Train models to predict PCA components
4. Compute SHAP values for each component

This reveals which parameters affect coverage of specific code regions.
"""
import ast
import json
from pathlib import Path
import numpy as np
from sklearn.decomposition import PCA
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import StandardScaler
import shap
import openpyxl

DATA_DIR = Path("data_VIS26/data_VIS26")
TUNERS = ["SymTuner", "CMA_ES", "Genetic", "SuccessiveHalving"]
PROGRAMS = ["gawk", "gcal", "grep"]

EXCLUDE_FROM_PARAMS = ["Iteration Coverage", "Accumulative Coverage"]
CATEGORICAL_PARAMS = {"seed-file", "seed_file", "seedfile"}
CATEGORICAL_THRESHOLD = 10
N_COMPONENTS = 10  # Number of PCA components to analyze


def load_parameters_raw(xlsx_path):
    """Load raw parameter values from xlsx."""
    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb.active

    param_names = []
    param_row_indices = []

    for row in range(3, sheet.max_row + 1):
        val = sheet.cell(row, 1).value
        if val and val not in EXCLUDE_FROM_PARAMS:
            param_names.append(val)
            param_row_indices.append(row)

    n_trials = sheet.max_column - 1
    raw_data = []

    for col in range(2, sheet.max_column + 1):
        trial_values = []
        for row in param_row_indices:
            val = sheet.cell(row, col).value
            trial_values.append(val)
        raw_data.append(trial_values)

    wb.close()
    return param_names, raw_data


def load_coverage_sets(coverage_path):
    """Load coverage sets - actual branch IDs covered by each trial."""
    coverage_sets = []
    with open(coverage_path, 'r') as f:
        for line in f:
            line = line.strip()
            if line:
                coverage_set = ast.literal_eval(line)
                coverage_sets.append(set(coverage_set))
    return coverage_sets


def create_coverage_matrix(coverage_sets):
    """Create binary coverage matrix (trials x branches)."""
    # Get all unique branches
    all_branches = set()
    for cs in coverage_sets:
        all_branches.update(cs)

    # Sort for consistent ordering
    branch_list = sorted(all_branches)
    branch_to_idx = {b: i for i, b in enumerate(branch_list)}

    # Create binary matrix
    n_trials = len(coverage_sets)
    n_branches = len(branch_list)

    matrix = np.zeros((n_trials, n_branches), dtype=np.float32)
    for i, cs in enumerate(coverage_sets):
        for branch in cs:
            matrix[i, branch_to_idx[branch]] = 1.0

    return matrix, branch_list


def detect_and_encode_params(param_names, raw_data):
    """Detect categorical params and encode all parameters."""
    n_trials = len(raw_data)
    categorical_info = {}
    string_params = set()

    # Detect categorical parameters
    for param_idx, param_name in enumerate(param_names):
        values = [trial[param_idx] for trial in raw_data if trial[param_idx] is not None]
        has_string = any(isinstance(v, str) for v in values)
        unique_vals = set(values)

        if has_string:
            string_params.add(param_name)

        is_boolean = unique_vals <= {True, False, 0, 1, 0.0, 1.0} and not has_string
        if is_boolean:
            continue

        is_categorical = (
            has_string or
            param_name.lower() in CATEGORICAL_PARAMS or
            len(unique_vals) <= CATEGORICAL_THRESHOLD
        )

        if is_categorical:
            sorted_vals = sorted([v for v in unique_vals],
                                key=lambda x: (isinstance(x, str), str(x) if isinstance(x, str) else x))
            categorical_info[param_name] = sorted_vals

    # Encode parameters
    encoded_columns = []
    encoded_param_names = []
    param_mapping = {}

    for param_idx, param_name in enumerate(param_names):
        values = [trial[param_idx] for trial in raw_data]

        if param_name in categorical_info:
            categories = categorical_info[param_name]
            param_encoded_names = []

            for cat in categories:
                encoded_name = f"{param_name}={cat}"
                encoded_param_names.append(encoded_name)
                param_encoded_names.append(encoded_name)
                column = np.array([1.0 if v == cat else 0.0 for v in values], dtype=np.float32)
                encoded_columns.append(column)

            param_mapping[param_name] = param_encoded_names
        else:
            encoded_param_names.append(param_name)
            param_mapping[param_name] = [param_name]

            column = np.zeros(n_trials, dtype=np.float32)
            for i, val in enumerate(values):
                if val is True:
                    column[i] = 1.0
                elif val is False:
                    column[i] = 0.0
                elif val is not None:
                    try:
                        column[i] = float(val)
                    except:
                        column[i] = 0.0
            encoded_columns.append(column)

    X = np.column_stack(encoded_columns) if encoded_columns else np.zeros((n_trials, 0))
    return X, encoded_param_names, param_mapping, param_names


def analyze_tuner_coverage_patterns(program, tuner):
    """Analyze coverage patterns for a tuner using PCA + SHAP."""
    base_path = DATA_DIR / program / tuner
    xlsx_path = base_path / "parameters.xlsx"
    coverage_path = base_path / "coverage_set"

    if not xlsx_path.exists() or not coverage_path.exists():
        return None

    print(f"  Loading {program}/{tuner}...")

    # Load data
    param_names, raw_data = load_parameters_raw(xlsx_path)
    coverage_sets = load_coverage_sets(coverage_path)

    # Align lengths
    min_len = min(len(raw_data), len(coverage_sets))
    raw_data = raw_data[:min_len]
    coverage_sets = coverage_sets[:min_len]

    print(f"    Trials: {min_len}, Parameters: {len(param_names)}")

    # Create coverage matrix
    coverage_matrix, branch_list = create_coverage_matrix(coverage_sets)
    print(f"    Branches: {len(branch_list)}")

    # Apply PCA
    print(f"    Applying PCA (n_components={N_COMPONENTS})...")

    # Standardize coverage matrix
    scaler = StandardScaler()
    coverage_scaled = scaler.fit_transform(coverage_matrix)

    pca = PCA(n_components=min(N_COMPONENTS, len(branch_list), min_len))
    coverage_pca = pca.fit_transform(coverage_scaled)

    explained_variance = pca.explained_variance_ratio_
    print(f"    Explained variance: {explained_variance[:5].sum():.1%} (first 5 components)")

    # Encode parameters
    X, encoded_param_names, param_mapping, original_param_names = detect_and_encode_params(param_names, raw_data)
    print(f"    Encoded features: {X.shape[1]}")

    # Train model and compute SHAP for each component
    print(f"    Computing SHAP values for each component...")

    component_results = []

    for comp_idx in range(pca.n_components_):
        y = coverage_pca[:, comp_idx]

        # Train model
        model = RandomForestRegressor(n_estimators=50, max_depth=8, random_state=42, n_jobs=-1)
        model.fit(X, y)

        # Compute SHAP
        explainer = shap.TreeExplainer(model)
        shap_values = explainer.shap_values(X)

        # Aggregate SHAP values to original parameters
        encoded_mean_shap = np.abs(shap_values).mean(axis=0)
        encoded_name_to_idx = {name: idx for idx, name in enumerate(encoded_param_names)}

        aggregated_importance = {}
        for original_param in original_param_names:
            encoded_names = param_mapping.get(original_param, [original_param])
            total_importance = sum(
                encoded_mean_shap[encoded_name_to_idx[enc_name]]
                for enc_name in encoded_names
                if enc_name in encoded_name_to_idx
            )
            aggregated_importance[original_param] = float(total_importance)

        # Sort by importance
        sorted_importance = sorted(aggregated_importance.items(), key=lambda x: x[1], reverse=True)

        # Get top branches contributing to this component
        component_loadings = pca.components_[comp_idx]
        top_branch_indices = np.argsort(np.abs(component_loadings))[-10:][::-1]
        top_branches = [
            {"branch_id": int(branch_list[i]), "loading": float(component_loadings[i])}
            for i in top_branch_indices
        ]

        component_results.append({
            "component_idx": comp_idx,
            "explained_variance": float(explained_variance[comp_idx]),
            "cumulative_variance": float(explained_variance[:comp_idx+1].sum()),
            "param_importance": [
                {"name": name, "importance": imp}
                for name, imp in sorted_importance[:20]
            ],
            "top_branches": top_branches,
            "component_values": coverage_pca[:, comp_idx].tolist()
        })

        if comp_idx < 3:
            top3 = [p[0] for p in sorted_importance[:3]]
            print(f"      PC{comp_idx+1} ({explained_variance[comp_idx]:.1%}): {', '.join(top3)}")

    # Also compute correlation between components and total coverage
    total_coverage = np.array([len(cs) for cs in coverage_sets])
    component_coverage_corr = [
        float(np.corrcoef(coverage_pca[:, i], total_coverage)[0, 1])
        for i in range(pca.n_components_)
    ]

    return {
        "n_trials": min_len,
        "n_branches": len(branch_list),
        "n_components": pca.n_components_,
        "total_variance_explained": float(explained_variance.sum()),
        "component_coverage_correlation": component_coverage_corr,
        "components": component_results,
        "coverage_stats": {
            "min": int(total_coverage.min()),
            "max": int(total_coverage.max()),
            "mean": float(total_coverage.mean())
        }
    }


def main():
    results = {}

    for program in PROGRAMS:
        results[program] = {}
        print(f"\n{'='*50}")
        print(f"Program: {program}")
        print('='*50)

        for tuner in TUNERS:
            result = analyze_tuner_coverage_patterns(program, tuner)
            if result:
                results[program][tuner] = result

    # Save results
    output_path = Path("public/data/coverage_pattern_data.json")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, 'w') as f:
        json.dump(results, f)

    print(f"\nSaved to {output_path}")
    print(f"File size: {output_path.stat().st_size / 1024:.1f} KB")

    # Summary
    print("\n" + "="*60)
    print("SUMMARY: Top parameters by component")
    print("="*60)

    for program in PROGRAMS:
        print(f"\n{program}:")
        for tuner in TUNERS:
            if tuner in results[program]:
                r = results[program][tuner]
                print(f"  {tuner}:")
                for comp in r["components"][:3]:
                    top_params = [p["name"] for p in comp["param_importance"][:3]]
                    corr = r["component_coverage_correlation"][comp["component_idx"]]
                    print(f"    PC{comp['component_idx']+1} ({comp['explained_variance']:.1%}, corr={corr:.2f}): {', '.join(top_params)}")


if __name__ == "__main__":
    main()
