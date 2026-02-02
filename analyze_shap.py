"""
SHAP Value Analysis for Tuner Comparison

Handles categorical parameters with one-hot encoding for accurate SHAP analysis.
"""
import openpyxl
import ast
import json
from pathlib import Path
import numpy as np

DATA_DIR = Path("data_VIS26/data_VIS26")
TUNERS = ["SymTuner", "CMA_ES", "Genetic", "SuccessiveHalving"]
PROGRAMS = ["gawk", "gcal", "grep"]

# These are OUTPUT values, not parameters - must exclude from SHAP analysis
EXCLUDE_FROM_PARAMS = ["Iteration Coverage", "Accumulative Coverage"]

# Parameters that should be treated as categorical even if numeric
CATEGORICAL_PARAMS = {"seed-file", "seed_file", "seedfile"}
CATEGORICAL_THRESHOLD = 10  # If unique values <= this, treat as categorical


def load_parameters(xlsx_path):
    """Load parameters from xlsx file (rows=params, cols=trials)

    Excludes output values like 'Iteration Coverage' and 'Accumulative Coverage'
    which are results, not input parameters.
    """
    wb = openpyxl.load_workbook(xlsx_path)
    sheet = wb.active

    # Get parameter names (column A, starting from row 3)
    # Also track which rows to include (excluding output values)
    all_param_names = []
    param_row_indices = []  # Original row indices (0-based from row 3)

    for row in range(3, sheet.max_row + 1):
        val = sheet.cell(row, 1).value
        if val:
            if val not in EXCLUDE_FROM_PARAMS:
                all_param_names.append(val)
                param_row_indices.append(row - 3)  # 0-based index

    # Get trial data (columns B onwards, row 2 has trial numbers)
    n_trials = sheet.max_column - 1
    n_params = len(all_param_names)

    # Create raw data matrix (trials x params) - keep original values
    raw_data = []
    for col in range(2, sheet.max_column + 1):
        trial_values = []
        for param_idx, orig_row_idx in enumerate(param_row_indices):
            row = 3 + orig_row_idx
            val = sheet.cell(row, col).value
            trial_values.append(val)
        raw_data.append(trial_values)

    wb.close()
    return all_param_names, raw_data


def detect_categorical_params(param_names, raw_data):
    """Detect which parameters should be treated as categorical.

    Returns:
        categorical_info: dict mapping param_name -> list of unique values (sorted)
        string_params: set of param names that have string values
    """
    n_trials = len(raw_data)
    categorical_info = {}  # param_name -> list of unique values (sorted)
    string_params = set()  # params that have string values

    for param_idx, param_name in enumerate(param_names):
        # Collect all values for this parameter
        values = []
        has_string = False
        for trial in raw_data:
            val = trial[param_idx]
            if val is not None:
                values.append(val)
                if isinstance(val, str):
                    has_string = True

        if has_string:
            string_params.add(param_name)

        unique_vals = set(values)

        # Check if boolean (only True/False or 0/1)
        is_boolean = unique_vals <= {True, False, 0, 1, 0.0, 1.0} and not has_string
        if is_boolean:
            continue  # Boolean is handled separately, not one-hot encoded

        # Check if categorical (string values OR low cardinality numeric OR specific params)
        is_categorical = (
            has_string or
            param_name.lower() in CATEGORICAL_PARAMS or
            len(unique_vals) <= CATEGORICAL_THRESHOLD
        )

        if is_categorical:
            # Sort unique values for consistent ordering
            # Strings come after numbers
            sorted_vals = sorted([v for v in unique_vals if v is not None],
                                  key=lambda x: (isinstance(x, str), str(x) if isinstance(x, str) else x))
            categorical_info[param_name] = sorted_vals
            type_str = "(string)" if has_string else ""
            print(f"    Categorical: {param_name} ({len(sorted_vals)} values) {type_str}")

    return categorical_info, string_params


def encode_parameters(param_names, raw_data, categorical_info):
    """Encode parameters with one-hot encoding for categorical features.

    Returns:
        encoded_param_names: list of encoded parameter names
        encoded_data: numpy array (trials x encoded_params)
        param_mapping: dict mapping original param -> list of encoded param names
    """
    n_trials = len(raw_data)
    encoded_columns = []
    encoded_param_names = []
    param_mapping = {}  # original_param -> [encoded_param_names]

    for param_idx, param_name in enumerate(param_names):
        values = [trial[param_idx] for trial in raw_data]

        if param_name in categorical_info:
            # One-hot encode categorical parameter
            categories = categorical_info[param_name]
            param_encoded_names = []

            for cat in categories:
                encoded_name = f"{param_name}={cat}"
                encoded_param_names.append(encoded_name)
                param_encoded_names.append(encoded_name)

                # Create one-hot column
                column = np.array([1.0 if v == cat else 0.0 for v in values], dtype=np.float32)
                encoded_columns.append(column)

            param_mapping[param_name] = param_encoded_names
        else:
            # Numeric or boolean - convert directly
            encoded_param_names.append(param_name)
            param_mapping[param_name] = [param_name]

            column = np.zeros(n_trials, dtype=np.float32)
            for i, val in enumerate(values):
                if val is True:
                    column[i] = 1.0
                elif val is False:
                    column[i] = 0.0
                elif val is not None:
                    try:
                        column[i] = float(val)
                    except:
                        column[i] = 0.0
            encoded_columns.append(column)

    encoded_data = np.column_stack(encoded_columns) if encoded_columns else np.zeros((n_trials, 0))
    return encoded_param_names, encoded_data, param_mapping


def load_coverage(coverage_path):
    """Load coverage sets and calculate coverage counts"""
    with open(coverage_path, 'r') as f:
        lines = f.readlines()

    coverages = []
    for line in lines:
        line = line.strip()
        if line:
            # Parse Python set literal
            coverage_set = ast.literal_eval(line)
            coverages.append(len(coverage_set))

    return np.array(coverages)


def analyze_tuner(program, tuner):
    """Load data and calculate SHAP values for a tuner"""
    base_path = DATA_DIR / program / tuner

    if not base_path.exists():
        return None

    xlsx_path = base_path / "parameters.xlsx"
    coverage_path = base_path / "coverage_set"

    if not xlsx_path.exists() or not coverage_path.exists():
        return None

    print(f"Loading {program}/{tuner}...")

    # Load raw data
    param_names, raw_data = load_parameters(xlsx_path)
    y = load_coverage(coverage_path)

    # Truncate if lengths don't match
    min_len = min(len(raw_data), len(y))
    raw_data = raw_data[:min_len]
    y = y[:min_len]

    print(f"  Parameters: {len(param_names)}, Trials: {len(y)}")
    print(f"  Coverage range: {y.min()} - {y.max()}, mean: {y.mean():.1f}")

    # Detect categorical parameters
    print(f"  Detecting categorical parameters...")
    categorical_info, string_params = detect_categorical_params(param_names, raw_data)

    # Encode parameters (one-hot for categorical)
    encoded_param_names, X, param_mapping = encode_parameters(param_names, raw_data, categorical_info)
    print(f"  Encoded features: {len(encoded_param_names)}")

    return {
        'param_names': param_names,
        'raw_data': raw_data,
        'encoded_param_names': encoded_param_names,
        'X': X,
        'y': y,
        'param_mapping': param_mapping,
        'categorical_info': categorical_info,
        'string_params': string_params,
        'program': program,
        'tuner': tuner
    }


def calculate_shap_values(data):
    """Calculate SHAP values using RandomForest with proper categorical handling"""
    from sklearn.ensemble import RandomForestRegressor
    import shap

    X = data['X']
    y = data['y']
    encoded_param_names = data['encoded_param_names']
    param_names = data['param_names']
    param_mapping = data['param_mapping']

    # Train model
    model = RandomForestRegressor(n_estimators=100, max_depth=10, random_state=42, n_jobs=-1)
    model.fit(X, y)

    # Calculate SHAP values
    explainer = shap.TreeExplainer(model)
    shap_values = explainer.shap_values(X)

    # Get mean absolute SHAP values for encoded features
    encoded_mean_shap = np.abs(shap_values).mean(axis=0)

    # Aggregate SHAP values back to original parameters
    # For categorical: sum of absolute SHAP values across all one-hot columns
    aggregated_importance = {}
    encoded_name_to_idx = {name: idx for idx, name in enumerate(encoded_param_names)}

    for original_param in param_names:
        encoded_names = param_mapping[original_param]
        total_importance = 0.0

        for enc_name in encoded_names:
            if enc_name in encoded_name_to_idx:
                idx = encoded_name_to_idx[enc_name]
                total_importance += encoded_mean_shap[idx]

        aggregated_importance[original_param] = total_importance

    # Create sorted importance ranking
    importance = [(param, aggregated_importance[param]) for param in param_names]
    importance.sort(key=lambda x: x[1], reverse=True)

    return {
        'shap_values': shap_values,
        'encoded_mean_shap': encoded_mean_shap,
        'aggregated_importance': aggregated_importance,
        'importance': importance,
        'model': model
    }


def analyze_all():
    """Analyze all tuners and save results"""
    results = {}

    for program in PROGRAMS:
        results[program] = {}
        print(f"\n{'='*50}")
        print(f"Program: {program}")
        print('='*50)

        for tuner in TUNERS:
            data = analyze_tuner(program, tuner)
            if data is None:
                continue

            print(f"\n  Calculating SHAP values for {tuner}...")
            shap_result = calculate_shap_values(data)

            print(f"  Top 10 influential parameters:")
            for i, (param, importance) in enumerate(shap_result['importance'][:10]):
                is_cat = "(categorical)" if param in data['categorical_info'] else ""
                print(f"    {i+1}. {param}: {importance:.2f} {is_cat}")

            results[program][tuner] = {
                'data': data,
                'shap': shap_result
            }

    return results


def export_decision_tree_data(results):
    """Export data in format suitable for Decision Tree visualization"""
    output = {}

    for program in PROGRAMS:
        output[program] = {}

        for tuner in TUNERS:
            if tuner not in results[program]:
                continue

            data = results[program][tuner]['data']
            shap_result = results[program][tuner]['shap']
            categorical_info = data['categorical_info']
            string_params = data['string_params']

            # Get parameter info with SHAP importance
            param_info = []
            for param, importance in shap_result['importance']:
                param_idx = data['param_names'].index(param)

                # Collect unique values from raw_data
                raw_values = [trial[param_idx] for trial in data['raw_data']]
                unique_vals = sorted(set(v for v in raw_values if v is not None),
                                     key=lambda x: (isinstance(x, str), str(x) if isinstance(x, str) else x))

                # Determine if boolean
                is_boolean = set(unique_vals) <= {True, False, 0, 1, 0.0, 1.0} and len(unique_vals) <= 2 and param not in string_params

                # Determine if string categorical
                is_string_categorical = param in string_params

                # Convert unique values for JSON
                if is_boolean:
                    unique_vals_out = [0.0, 1.0]
                elif is_string_categorical:
                    # Keep string values as-is for string categorical
                    unique_vals_out = [str(v) for v in unique_vals]
                else:
                    unique_vals_out = []
                    for v in unique_vals:
                        if isinstance(v, bool):
                            unique_vals_out.append(1.0 if v else 0.0)
                        elif isinstance(v, (int, float)):
                            unique_vals_out.append(float(v))
                        elif isinstance(v, str):
                            # String in non-string param (shouldn't happen but handle it)
                            unique_vals_out.append(v)

                param_info.append({
                    'name': param,
                    'importance': float(importance),
                    'unique_values': unique_vals_out,
                    'is_boolean': is_boolean,
                    'is_categorical': param in categorical_info,
                    'is_string': is_string_categorical
                })

            # Prepare trial data with coverage
            trials = []
            for i in range(len(data['y'])):
                trial_params = {}
                for j, param_name in enumerate(data['param_names']):
                    val = data['raw_data'][i][j]

                    # Convert to appropriate type
                    if val is True:
                        trial_params[param_name] = True
                    elif val is False:
                        trial_params[param_name] = False
                    elif val is None:
                        trial_params[param_name] = None
                    elif isinstance(val, str):
                        # Keep string values for string categorical params
                        trial_params[param_name] = val
                    else:
                        try:
                            trial_params[param_name] = float(val)
                        except:
                            trial_params[param_name] = str(val) if val else None

                trials.append({
                    'trial_id': i + 1,
                    'coverage': int(data['y'][i]),
                    'parameters': trial_params
                })

            output[program][tuner] = {
                'param_importance': param_info,
                'trials': trials,
                'stats': {
                    'total_trials': len(data['y']),
                    'min_coverage': int(data['y'].min()),
                    'max_coverage': int(data['y'].max()),
                    'mean_coverage': float(data['y'].mean())
                }
            }

    return output


if __name__ == "__main__":
    results = analyze_all()

    # Save top parameters summary
    print("\n\n" + "="*60)
    print("SUMMARY: Top 5 Parameters by Tuner")
    print("="*60)

    for program in PROGRAMS:
        print(f"\n{program}:")
        for tuner in TUNERS:
            if tuner in results[program]:
                top5 = results[program][tuner]['shap']['importance'][:5]
                top_names = [p[0] for p in top5]
                print(f"  {tuner}: {', '.join(top_names)}")

    # Export for visualization
    print("\n\nExporting data for visualization...")
    export_data = export_decision_tree_data(results)

    # Save to JSON
    output_path = Path("public/data/decision_tree_data.json")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, 'w') as f:
        json.dump(export_data, f)

    print(f"Saved to {output_path}")
    print(f"File size: {output_path.stat().st_size / 1024 / 1024:.2f} MB")
